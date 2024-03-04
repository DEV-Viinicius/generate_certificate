from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

def generate_certificates(filepath):
    try:
        workbook_alunos = load_workbook(filepath)
        sheet_alunos = workbook_alunos.active

        for indice, linhas in enumerate(sheet_alunos.iter_rows(min_row=2), start=1):
            nome_curso = linhas[0].value 
            nome_participante = linhas[1].value
            tipo = linhas[2].value
            data_inicio = linhas[3].value
            data_fim = linhas[4].value
            carga_horaria = linhas[5].value
            data_emissao = linhas[6].value 

            fonte_nome = ImageFont.truetype('./tahomabd.ttf', 90)
            fonte_geral = ImageFont.truetype('./tahoma.ttf', 80)
            fonte_data = ImageFont.truetype('./tahoma.ttf', 55)

            image = Image.open('./certificado_padrao.jpg')
            desenhar = ImageDraw.Draw(image)

            desenhar.text((1020, 827), nome_participante, fill='black', font=fonte_nome)
            desenhar.text((1060, 950), nome_curso, fill='black', font=fonte_geral)
            desenhar.text((1435, 1065), tipo, fill='black', font=fonte_geral)
            desenhar.text((1480, 1182), str(carga_horaria), fill='black', font=fonte_geral)
            desenhar.text((750, 1770), data_inicio, fill='black', font=fonte_data)
            desenhar.text((750, 1930), data_fim, fill='black', font=fonte_data)
            desenhar.text((2220, 1930), data_emissao, fill='black', font=fonte_data)

            image.save(f'./{indice} {nome_participante} certificado.png')

        return True, "Certificados gerados com sucesso!"
    except Exception as e:
        return False, f"Ocorreu um erro ao gerar os certificados: {str(e)}"
